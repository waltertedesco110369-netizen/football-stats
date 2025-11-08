"""
Script per copiare il foglio da un file Excel convertito al template
Sostituisce il foglio IMPORT_TEMP mantenendo il foglio DATI protetto
"""
import pandas as pd
from openpyxl import load_workbook
import sys
import os

def copy_excel_to_template(excel_file_path, template_path="Import/Database/File/template_calcio_base.xlsx"):
    """
    Copia il primo foglio del file Excel convertito nel template
    Sostituisce il foglio IMPORT_TEMP mantenendo il foglio DATI protetto
    """
    
    if not os.path.exists(excel_file_path):
        print(f"Errore: file non trovato: {excel_file_path}")
        return False
    
    if not os.path.exists(template_path):
        print(f"Errore: template non trovato: {template_path}")
        print("Esegui prima: python create_template_excel.py")
        return False
    
    try:
        # Leggi il file Excel convertito (primo foglio)
        df_convertito = pd.read_excel(excel_file_path, sheet_name=0, header=None)
        
        # Carica il template
        wb = load_workbook(template_path)
        
        # Rimuovi il foglio IMPORT_TEMP se esiste
        if 'IMPORT_TEMP' in wb.sheetnames:
            wb.remove(wb['IMPORT_TEMP'])
        
        # Crea nuovo foglio IMPORT_TEMP con i dati convertiti
        with pd.ExcelWriter(template_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            writer.book = wb
            writer.sheets = {name: wb[name] for name in wb.sheetnames}
            df_convertito.to_excel(writer, sheet_name='IMPORT_TEMP', index=False, header=False)
        
        # Salva il template (mantiene il foglio DATI protetto)
        wb.save(template_path)
        
        print(f"✓ File copiato con successo!")
        print(f"  - File convertito: {os.path.basename(excel_file_path)}")
        print(f"  - Template aggiornato: {template_path}")
        print(f"  - Foglio IMPORT_TEMP sostituito")
        print(f"  - Foglio DATI protetto mantenuto ✓")
        print(f"\nOra esegui: python map_excel_to_template.py")
        return True
        
    except Exception as e:
        print(f"Errore durante la copia: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python copy_excel_to_template.py <percorso_file_excel_convertito>")
        print("\nEsempio:")
        print("  python copy_excel_to_template.py \"Import/Database/File/calcio base per data (2).xlsx\"")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    copy_excel_to_template(excel_file)

