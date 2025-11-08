"""
Script per creare il template Excel con foglio protetto
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter

def create_template_excel(output_path="Import/Database/File/template_calcio_base.xlsx"):
    """Crea un template Excel con foglio DATI protetto e foglio IMPORT_TEMP"""
    
    # Colonne compatibili con il database (mappatura new_leagues)
    colonne_base = [
        'League', 'Season', 'Date', 'Time', 'Home', 'Away',
        'HG', 'AG', 'Res', 'Live'
    ]
    
    colonne_quote = [
        'quota_1', 'quota_X', 'quota_2', 'H', 'H1', 'HX', 'H2',
        '1X', 'X2', '12',
        'uo_1_5_u', 'uo_1_5_o', 'uo_2_5_u', 'uo_2_5_o', 'uo_3_5_u', 'uo_3_5_o',
        'G', 'NO_G', 'C_SI', 'C_NO', 'O_SI', 'O_NO'
    ]
    
    colonne_extra = ['pal', 'avv']
    
    tutte_colonne = colonne_base + colonne_quote + colonne_extra
    
    # Crea DataFrame vuoto con tutte le colonne
    df_dati = pd.DataFrame(columns=tutte_colonne)
    
    # Crea DataFrame vuoto per IMPORT_TEMP
    df_temp = pd.DataFrame()
    
    # Salva in Excel con due fogli
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_dati.to_excel(writer, sheet_name='DATI', index=False)
        df_temp.to_excel(writer, sheet_name='IMPORT_TEMP', index=False)
    
    # Carica il workbook per proteggere il foglio DATI
    wb = load_workbook(output_path)
    ws_dati = wb['DATI']
    
    # Proteggi il foglio DATI (le celle sono editabili solo via script)
    ws_dati.protection.sheet = True
    ws_dati.protection.enable()
    
    # Larghezza colonne ottimale
    for idx, col in enumerate(tutte_colonne, start=1):
        ws_dati.column_dimensions[get_column_letter(idx)].width = 15
    
    # Salva il file
    wb.save(output_path)
    print(f"Template creato: {output_path}")
    print(f"- Foglio 'DATI': PROTETTO con {len(tutte_colonne)} colonne")
    print(f"- Foglio 'IMPORT_TEMP': non protetto per importazione")

if __name__ == "__main__":
    import os
    os.makedirs("Import/Database/File", exist_ok=True)
    create_template_excel()

