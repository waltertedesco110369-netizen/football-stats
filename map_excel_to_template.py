"""
Script per mappare i dati dal foglio IMPORT_TEMP al foglio DATI protetto
Gestisce merge intelligente: aggiunge nuovi record o aggiorna solo campi vuoti
"""
import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter

def estrai_data_italiana(text):
    """Estrae data da testo tipo 'sabato, 1 novembre' e ritorna (data, stagione)"""
    giorni_it = r"lunedì|lunedi|martedì|martedi|mercoledì|mercoledi|giovedì|giovedi|venerdì|venerdi|sabato|domenica"
    mesi_it = "gennaio|febbraio|marzo|aprile|maggio|giugno|luglio|agosto|settembre|ottobre|novembre|dicembre"
    date_header_re = re.compile(fr"^(?:{giorni_it})\s*,\s*(\d{{1,2}})\s+({mesi_it})$", re.IGNORECASE)
    
    month_map = {
        'gennaio': 1, 'febbraio': 2, 'marzo': 3, 'aprile': 4, 'maggio': 5, 'giugno': 6,
        'luglio': 7, 'agosto': 8, 'settembre': 9, 'ottobre': 10, 'novembre': 11, 'dicembre': 12
    }
    
    current_year = datetime.now().year
    
    if isinstance(text, str):
        match = date_header_re.match(text.strip())
        if match:
            dd = int(match.group(1))
            mm_name = match.group(2).lower()
            mm = month_map.get(mm_name)
            if mm:
                data = f"{current_year}-{mm:02d}-{dd:02d}"
                stagione = str(current_year)
                return data, stagione
    return None, None

def normalizza_valore(val):
    """Normalizza un valore per Excel"""
    if pd.isna(val) or val == '' or val == 'nan' or val == 'None':
        return None
    val_str = str(val).strip()
    if not val_str:
        return None
    # Converti virgola in punto per decimali
    val_str = val_str.replace(',', '.')
    try:
        # Prova come float
        return float(val_str)
    except:
        try:
            # Prova come int
            return int(val_str)
        except:
            # Resta stringa
            return val_str

def map_excel_to_template(template_path="Import/Database/File/template_calcio_base.xlsx"):
    """
    Legge i dati dal foglio IMPORT_TEMP e li mappa nel foglio DATI protetto
    Fa merge intelligente: conserva dati esistenti, aggiorna solo campi vuoti
    """
    
    try:
        # Carica il template
        wb = load_workbook(template_path)
        
        # Leggi foglio IMPORT_TEMP (dati convertiti)
        try:
            df_temp = pd.read_excel(template_path, sheet_name='IMPORT_TEMP', header=None)
        except:
            print("Errore: foglio IMPORT_TEMP vuoto o non trovato")
            return False
        
        if df_temp.empty:
            print("Foglio IMPORT_TEMP vuoto")
            return False
        
        # Leggi foglio DATI esistente (dati già mappati)
        try:
            df_dati_esistenti = pd.read_excel(template_path, sheet_name='DATI')
        except:
            df_dati_esistenti = pd.DataFrame()
        
        # Trova riga intestazioni (cerca "Ora", "Manif", "Squadra 1", ecc.)
        riga_intestazioni = None
        for idx in range(min(5, len(df_temp))):
            row_values = [str(v).strip() if pd.notna(v) else '' for v in df_temp.iloc[idx]]
            if any('ora' in str(v).lower() or 'manif' in str(v).lower() or 'squadra' in str(v).lower() 
                   for v in row_values[:10]):
                riga_intestazioni = idx
                break
        
        if riga_intestazioni is None:
            print("Errore: intestazioni non trovate nel foglio IMPORT_TEMP")
            print("Provo a cercare nella riga 2 (standard)...")
            if len(df_temp) > 2:
                riga_intestazioni = 2
            else:
                return False
        
        # Leggi con intestazioni dalla riga trovata
        df_temp = pd.read_excel(template_path, sheet_name='IMPORT_TEMP', header=riga_intestazioni)
        
        # Cerca anche riga data (tipo "sabato, 1 novembre")
        riga_data = None
        df_temp_raw = pd.read_excel(template_path, sheet_name='IMPORT_TEMP', header=None)
        for idx in range(min(10, len(df_temp_raw))):
            for val in df_temp_raw.iloc[idx]:
                if pd.notna(val):
                    data, stagione = estrai_data_italiana(str(val))
                    if data:
                        riga_data = idx
                        break
            if riga_data is not None:
                break
        
        # Mappa colonne del file convertito
        # Cerca colonne per nome (case-insensitive)
        def trova_colonna(df, possibili_nomi):
            for nome in possibili_nomi:
                for col in df.columns:
                    if nome.lower() in str(col).lower():
                        return col
            return None
        
        # Estrai dati da ogni riga
        nuovi_dati = []
        current_date = None
        current_season = None
        
        for idx, row in df_temp.iterrows():
            # Skip righe vuote
            if row.isna().all():
                continue
            
            # Cerca data italiana nella riga
            for val in row:
                if pd.notna(val):
                    data, stagione = estrai_data_italiana(str(val))
                    if data:
                        current_date = data
                        current_season = stagione
                        break
            
            # Estrai dati base
            ora_col = trova_colonna(df_temp, ['Ora', 'Time', 'ora', 'time'])
            manif_col = trova_colonna(df_temp, ['Manif', 'Manif.', 'League', 'league'])
            pal_col = trova_colonna(df_temp, ['Pal', 'Pal.', 'pal'])
            avv_col = trova_colonna(df_temp, ['Avv', 'Avv.', 'avv'])
            squadra1_col = trova_colonna(df_temp, ['Squadra 1', 'Home', 'home', 'Squadra1'])
            squadra2_col = trova_colonna(df_temp, ['Squadra 2', 'Away', 'away', 'Squadra2'])
            live_col = trova_colonna(df_temp, ['LIVE', 'Live', 'live', 'L I V E'])
            
            ora = str(row[ora_col]).strip() if ora_col and pd.notna(row[ora_col]) else ''
            manif = str(row[manif_col]).strip() if manif_col and pd.notna(row[manif_col]) else ''
            pal = str(row[pal_col]).strip() if pal_col and pd.notna(row[pal_col]) else ''
            avv = str(row[avv_col]).strip() if avv_col and pd.notna(row[avv_col]) else ''
            squadra1 = str(row[squadra1_col]).strip() if squadra1_col and pd.notna(row[squadra1_col]) else ''
            squadra2 = str(row[squadra2_col]).strip() if squadra2_col and pd.notna(row[squadra2_col]) else ''
            live = 'LIVE' if (live_col and pd.notna(row[live_col]) and 'LIVE' in str(row[live_col]).upper()) else ''
            
            # Skip se mancano dati essenziali
            if not ora or not manif or not squadra1 or not squadra2:
                continue
            
            # Pulisci manif (rimuovi numeri finali)
            manif_clean = re.sub(r'\d+$', '', manif).upper().strip()
            
            # Estrai quote basandosi sulla posizione delle colonne (dopo LIVE)
            # Mappatura colonne dal file convertito:
            # Col 0: Ora, Col 1: Manif, Col 2: Pal, Col 3: Avv, Col 4: Squadra 1, Col 5: Squadra 2, Col 6: LIVE
            # Col 7+: quote (quota_1, quota_X, quota_2, H, H1, HX, H2, 1X, X2, 12, uo_1_5_u, uo_1_5_o, ecc.)
            
            quote = {}
            # Mappa colonne per indice (più affidabile dei nomi)
            colonne_quote_standard = [
                'quota_1', 'quota_X', 'quota_2', 'H', 'H1', 'HX', 'H2',
                '1X', 'X2', '12',
                'uo_1_5_u', 'uo_1_5_o', 'uo_2_5_u', 'uo_2_5_o', 'uo_3_5_u', 'uo_3_5_o',
                'G', 'NO_G', 'C_SI', 'C_NO', 'O_SI', 'O_NO'
            ]
            
            # Cerca colonne dopo LIVE (dalla posizione 7 in poi)
            colonne_ordine = list(df_temp.columns)
            start_idx = 7  # Dopo LIVE (colonne 0-6 sono: Ora, Manif, Pal, Avv, Squadra1, Squadra2, LIVE)
            
            for i, col_quote in enumerate(colonne_quote_standard):
                col_idx = start_idx + i
                if col_idx < len(colonne_ordine):
                    col_name = colonne_ordine[col_idx]
                    val = row[col_name]
                    if pd.notna(val):
                        val_norm = normalizza_valore(val)
                        if val_norm is not None:
                            quote[col_quote] = val_norm
            
            # Crea record
            record = {
                'League': manif_clean,
                'Season': current_season or str(datetime.now().year),
                'Date': current_date or '',
                'Time': ora,
                'Home': squadra1,
                'Away': squadra2,
                'Live': live,
                'pal': pal,
                'avv': avv,
                'HG': None,
                'AG': None,
                'Res': None,
                **quote  # Aggiungi tutte le quote trovate
            }
            
            nuovi_dati.append(record)
        
        if not nuovi_dati:
            print("Nessun dato valido trovato nel foglio IMPORT_TEMP")
            return False
        
        # Crea DataFrame con nuovi dati
        df_nuovi = pd.DataFrame(nuovi_dati)
        
        # Merge con dati esistenti
        if not df_dati_esistenti.empty:
            # Match per: League + Season + Date + Time + Home + Away
            colonne_match = ['League', 'Season', 'Date', 'Time', 'Home', 'Away']
            
            # Crea DataFrame risultato
            df_risultato = df_dati_esistenti.copy()
            
            for idx_nuovo, row_nuovo in df_nuovi.iterrows():
                # Cerca match (tutti i campi di match devono corrispondere)
                match_found = False
                for idx_esistente, row_esistente in df_risultato.iterrows():
                    # Verifica match: tutti i campi di match devono corrispondere
                    match = True
                    for col in colonne_match:
                        val_esistente = str(row_esistente[col]).strip() if pd.notna(row_esistente[col]) else ''
                        val_nuovo = str(row_nuovo[col]).strip() if pd.notna(row_nuovo[col]) else ''
                        if val_esistente != val_nuovo:
                            match = False
                            break
                    
                    if match:
                        match_found = True
                        # Aggiorna solo campi vuoti (non sovrascrive dati esistenti)
                        for col in df_risultato.columns:
                            if col in row_nuovo:
                                val_esistente = row_esistente[col]
                                val_nuovo = row_nuovo[col]
                                
                                # Aggiorna solo se il campo esistente è vuoto
                                is_empty = (
                                    pd.isna(val_esistente) or 
                                    str(val_esistente).strip() == '' or
                                    str(val_esistente).strip() == 'nan' or
                                    str(val_esistente).strip() == 'None'
                                )
                                
                                if is_empty and pd.notna(val_nuovo):
                                    df_risultato.at[idx_esistente, col] = val_nuovo
                        break
                
                # Se non c'è match, aggiungi nuovo record
                if not match_found:
                    df_risultato = pd.concat([df_risultato, pd.DataFrame([row_nuovo])], ignore_index=True)
        else:
            df_risultato = df_nuovi
        
        # Sblocca temporaneamente il foglio DATI per scrivere
        ws_dati = wb['DATI']
        ws_dati.protection.sheet = False
        
        # Salva nel foglio DATI
        with pd.ExcelWriter(template_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            writer.book = wb
            writer.sheets = {name: wb[name] for name in wb.sheetnames}
            df_risultato.to_excel(writer, sheet_name='DATI', index=False)
        
        # Riproteggi il foglio DATI
        ws_dati = wb['DATI']
        ws_dati.protection.sheet = True
        ws_dati.protection.enable()
        
        # Salva
        wb.save(template_path)
        
        print(f"✓ Dati mappati con successo!")
        print(f"  - Nuovi record: {len(nuovi_dati)}")
        print(f"  - Totale record nel foglio DATI: {len(df_risultato)}")
        return True
        
    except Exception as e:
        print(f"Errore durante la mappatura: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    map_excel_to_template()

