"""
Script UNICO che fa tutto automaticamente:
1. Copia il file Excel convertito nel template
2. Mappa i dati nel foglio DATI protetto
3. Opzionalmente importa nel database
"""
import pandas as pd
from openpyxl import load_workbook
import sys
import os
import time
from pathlib import Path

def importa_excel_automatico(excel_file_path, importa_db=False):
    """
    Processo automatico completo:
    1. Copia file Excel nel template
    2. Mappa dati nel foglio DATI
    3. Opzionalmente importa nel database
    """
    
    template_path = "Import/Database/File/template_calcio_base.xlsx"
    
    print("="*80)
    print("IMPORTAZIONE AUTOMATICA EXCEL CONVERTITO")
    print("="*80)
    
    # Verifica file
    if not os.path.exists(excel_file_path):
        print(f"❌ ERRORE: File non trovato: {excel_file_path}")
        return False
    
    if not os.path.exists(template_path):
        print(f"❌ ERRORE: Template non trovato. Creo il template...")
        from create_template_excel import create_template_excel
        create_template_excel(template_path)
    
    # PASSO 1: Copia il file nel template (chiudi se aperto)
    print(f"\n[1/3] Copio file Excel nel template...")
    try:
        # Prova a chiudere il file se aperto (attendi se bloccato)
        max_tentativi = 5
        for tentativo in range(max_tentativi):
            try:
                # Leggi il file Excel convertito
                df_convertito = pd.read_excel(excel_file_path, sheet_name=0, header=None)
                
                # Carica template
                wb = load_workbook(template_path)
                
                # Rimuovi IMPORT_TEMP se esiste
                if 'IMPORT_TEMP' in wb.sheetnames:
                    wb.remove(wb['IMPORT_TEMP'])
                
                # Salva il template prima (chiude eventuali lock)
                wb.save(template_path)
                
                # Crea nuovo IMPORT_TEMP
                with pd.ExcelWriter(template_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    writer.book = wb
                    writer.sheets = {name: wb[name] for name in wb.sheetnames}
                    df_convertito.to_excel(writer, sheet_name='IMPORT_TEMP', index=False, header=False)
                
                wb.save(template_path)
                print(f"   ✓ File copiato con successo!")
                break
                
            except PermissionError:
                if tentativo < max_tentativi - 1:
                    print(f"   ⚠ File aperto. Chiudilo e riprovo tra 2 secondi... (tentativo {tentativo+1}/{max_tentativi})")
                    time.sleep(2)
                else:
                    print(f"   ❌ ERRORE: File template aperto. Chiudilo e riprova.")
                    return False
            except Exception as e:
                print(f"   ❌ ERRORE: {e}")
                return False
    except Exception as e:
        print(f"   ❌ ERRORE durante la copia: {e}")
        return False
    
    # PASSO 2: Mappa i dati nel foglio DATI
    print(f"\n[2/3] Mappo i dati nel foglio DATI protetto...")
    try:
        from map_excel_to_template import map_excel_to_template
        if map_excel_to_template(template_path):
            print(f"   ✓ Dati mappati con successo!")
        else:
            print(f"   ❌ ERRORE durante la mappatura")
            return False
    except Exception as e:
        print(f"   ❌ ERRORE durante la mappatura: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # PASSO 3: Importa nel database (opzionale)
    if importa_db:
        print(f"\n[3/3] Importo nel database...")
        try:
            from database import FootballDatabase
            db = FootballDatabase(environment="test")
            if db.import_excel_file(template_path):
                print(f"   ✓ Importato nel database con successo!")
            else:
                print(f"   ❌ ERRORE durante l'importazione")
                return False
        except Exception as e:
            print(f"   ❌ ERRORE durante l'importazione: {e}")
            return False
    else:
        print(f"\n[3/3] Salto importazione database (usa -import per importare)")
    
    print("\n" + "="*80)
    print("✓ COMPLETATO CON SUCCESSO!")
    print("="*80)
    print(f"\nFile template pronto: {template_path}")
    print(f"Foglio DATI protetto aggiornato con i dati mappati.")
    if not importa_db:
        print(f"\nPer importare nel database, esegui:")
        print(f"  python importa_excel_convertito.py \"{excel_file_path}\" -import")
    
    return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python importa_excel_convertito.py <file_excel_convertito> [-import]")
        print("\nEsempi:")
        print("  python importa_excel_convertito.py \"Import/Database/File/calcio base per data (2).xlsx\"")
        print("  python importa_excel_convertito.py \"Import/Database/File/calcio base per data (2).xlsx\" -import")
        print("\n  -import: importa automaticamente nel database dopo la mappatura")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    importa_db = "-import" in sys.argv
    
    importa_excel_automatico(excel_file, importa_db)

